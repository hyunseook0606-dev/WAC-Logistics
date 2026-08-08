# -*- coding: utf-8 -*-
"""Build WAC Air Quotation Simulator (.xlsx) — clean one-page UI."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation

OUT = Path(__file__).resolve().parent / "WAC_Air_Quotation_Simulator.xlsx"

NAVY = "1A2A3A"
ORANGE = "F05023"
LIGHT = "F8FAFC"
INPUT_BG = "FFF7ED"
YELLOW = "FEF3C7"
GREEN_BG = "DCFCE7"
GREEN = "166534"
MUTED = "64748B"
LINE = "E2E8F0"
WHITE = "FFFFFF"
SLATE = "334155"

thin = Border(
    left=Side(style="thin", color=LINE),
    right=Side(style="thin", color=LINE),
    top=Side(style="thin", color=LINE),
    bottom=Side(style="thin", color=LINE),
)


def fill(c: str) -> PatternFill:
    return PatternFill("solid", fgColor=c)


def fnt(bold=False, size=11, color=NAVY):
    return Font(name="Calibri", bold=bold, size=size, color=color)


def unlock(cell):
    cell.protection = Protection(locked=False)


def lock(cell):
    cell.protection = Protection(locked=True)


def set_widths(ws, d):
    for k, v in d.items():
        ws.column_dimensions[k].width = v


def paint_merge(ws, cell_range, bg):
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = fill(bg)
            cell.border = thin
            lock(cell)


def header_bar(ws, cell_range, text, bg=NAVY):
    top = cell_range.split(":")[0]
    ws[top] = text
    ws[top].font = fnt(bold=True, size=11, color=WHITE)
    ws[top].alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(cell_range)
    paint_merge(ws, cell_range, bg)


def label_cell(cell, text):
    cell.value = text
    cell.font = fnt(bold=True, size=10, color=MUTED)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border = thin
    cell.fill = fill(WHITE)
    lock(cell)


def input_cell(cell, value=None, fmt=None):
    if value is not None:
        cell.value = value
    cell.fill = fill(INPUT_BG)
    cell.border = thin
    cell.font = fnt(bold=True, size=12)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt
    unlock(cell)


def value_cell(cell, formula, fmt=None, emphasize=False):
    cell.value = formula
    cell.fill = fill(GREEN_BG if emphasize else LIGHT)
    cell.border = thin
    cell.font = fnt(bold=emphasize, size=12 if emphasize else 11, color=GREEN if emphasize else NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        cell.number_format = fmt
    lock(cell)


def build_master(ws):
    ws.sheet_view.showGridLines = False
    set_widths(ws, {c: w for c, w in zip("ABCDEFGH", [22, 12, 12, 12, 12, 12, 12, 14])})

    ws["A1"] = "Master_DB  (Backend)"
    ws["A1"].font = fnt(bold=True, size=16)
    ws.merge_cells("A1:H1")
    ws["A2"] = "노란 칸만 수정 · Quote 단가는 전부 이 시트 참조 (하드코딩 금지)"
    ws["A2"].font = fnt(size=10, color=MUTED)
    ws.merge_cells("A2:H2")

    header_bar(ws, "A4:B4", "PARAMETERS")
    for r, name, val, note in (
        (5, "Vol factor (kg/CBM)", 167, "CBM x 167 = 부피중량"),
        (6, "CBM divisor (cm)", 1_000_000, "L x W x H x Qty / 1,000,000"),
    ):
        label_cell(ws[f"A{r}"], name)
        input_cell(ws[f"B{r}"], val, "0")
        ws[f"B{r}"].fill = fill(YELLOW)
        ws[f"C{r}"] = note
        ws[f"C{r}"].font = fnt(size=9, color=MUTED)
        lock(ws[f"C{r}"])

    header_bar(ws, "A8:B8", "WEIGHT BREAK (kg)")
    for i, (name, val) in enumerate(
        [("WB_45", 45), ("WB_100", 100), ("WB_500", 500), ("WB_1000", 1000)]
    ):
        r = 9 + i
        label_cell(ws[f"A{r}"], name)
        input_cell(ws[f"B{r}"], val, "0")
        ws[f"B{r}"].fill = fill(YELLOW)

    ws["A14"] = "1. Air freight & surcharge"
    ws["A14"].font = fnt(bold=True, size=12, color=ORANGE)
    ws.merge_cells("A14:H14")

    for c, h in enumerate(
        ["ROUTE", "MIN", "+45", "+100", "+500", "+1000", "FSC/kg", "SSC/kg"], 1
    ):
        cell = ws.cell(15, c, h)
        cell.fill = fill(NAVY)
        cell.font = fnt(bold=True, size=10, color=WHITE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin
        lock(cell)

    for i, row in enumerate(
        [
            ("ICN-HKG", 50, 4.5, 3.8, 3.2, 2.8, 0.6, 0.15),
            ("ICN-LAX", 90, 8.5, 7.2, 6.5, 5.8, 1.2, 0.25),
        ]
    ):
        r = 16 + i
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, v)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if c == 1:
                cell.font = fnt(bold=True)
                cell.fill = fill(INPUT_BG)
            else:
                cell.number_format = "0.00"
                cell.fill = fill(YELLOW)
            unlock(cell)

    ws["A19"] = "2. Local charges"
    ws["A19"].font = fnt(bold=True, size=12, color=ORANGE)

    for c, h in enumerate(["Charge Item", "Unit", "Rate", "MIN", "Note"], 1):
        cell = ws.cell(20, c, h)
        cell.fill = fill(NAVY)
        cell.font = fnt(bold=True, size=10, color=WHITE)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin
        lock(cell)

    for i, row in enumerate(
        [
            ("Handling Fee", "Per Shipment", 30, 30, "건당"),
            ("Doc Fee", "Per BL", 25, 25, "BL당"),
            ("Trucking", "Per CBM", 15, 80, "부피, MIN"),
        ]
    ):
        r = 21 + i
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, v)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")
            if c in (3, 4):
                cell.number_format = "0.00"
                cell.fill = fill(YELLOW)
            else:
                cell.fill = fill(LIGHT)
            unlock(cell)

    ws["A25"] = "행 16-17 / 21-23 삭제 금지 (Quote lookup 범위)"
    ws["A25"].font = fnt(size=9, color=MUTED)
    ws.freeze_panes = "A4"


def build_quote(ws):
    ws.sheet_view.showGridLines = False
    set_widths(
        ws,
        {
            "A": 2,
            "B": 18,
            "C": 14,
            "D": 2,
            "E": 18,
            "F": 14,
            "G": 2,
            "H": 14,
            "I": 14,
        },
    )
    for r in range(1, 30):
        ws.row_dimensions[r].height = 22
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[6].height = 34
    ws.row_dimensions[7].height = 34
    ws.row_dimensions[8].height = 34

    ws["B2"] = "항공 가견적 시뮬레이터"
    ws["B2"].font = fnt(bold=True, size=18)
    ws.merge_cells("B2:I2")
    lock(ws["B2"])

    ws["B3"] = "주황 칸만 입력 · 단가는 Master_DB 자동 조회 · 가견적(APPX), 최종 INV 아님"
    ws["B3"].font = fnt(size=9, color=MUTED)
    ws.merge_cells("B3:I3")
    lock(ws["B3"])

    # 1. Input
    header_bar(ws, "B5:C5", "1. 화물 입력", ORANGE)
    for r, lab, val, fmt in (
        (6, "Route", "ICN-HKG", "@"),
        (7, "L (cm)", 110, "0"),
        (8, "W (cm)", 110, "0"),
        (9, "H (cm)", 150, "0"),
        (10, "Qty (pcs)", 3, "0"),
        (11, "Gross (KG)", 400, "0.00"),
        (12, "B/L 수", 1, "0"),
    ):
        label_cell(ws[f"B{r}"], lab)
        input_cell(ws[f"C{r}"], val, fmt)
        ws.row_dimensions[r].height = 24

    dv = DataValidation(
        type="list",
        formula1='"ICN-HKG,ICN-LAX"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Route",
        error="Master_DB ROUTE만 선택",
    )
    dv.add("C6")
    ws.add_data_validation(dv)

    # 2. Auto calc
    header_bar(ws, "E5:F5", "2. 자동 계산", NAVY)
    for r, lab, formula, fmt, emp in (
        (6, "CBM", '=IFERROR(IF(COUNTA(C7:C10)<4,"",C7*C8*C9*C10/Master_DB!$B$6),"")', "0.000", False),
        (7, "부피중량 (KG)", '=IFERROR(IF(F6="","",F6*Master_DB!$B$5),"")', "0.00", False),
        (8, "C.W. (KG)", '=IFERROR(IF(OR(C11="",F7=""),"",MAX(C11,F7)),"")', "0.00", True),
        (
            9,
            "Weight Break",
            '=IFERROR(IF(F8="","",IF(F8>=Master_DB!$B$12,"+1000",IF(F8>=Master_DB!$B$11,"+500",IF(F8>=Master_DB!$B$10,"+100","+45")))),"")',
            "@",
            False,
        ),
        (
            10,
            "Air Rate ($/kg)",
            '=IFERROR(IF(OR(C6="",F8=""),"",IF(F8>=Master_DB!$B$12,VLOOKUP(C6,Master_DB!$A$16:$H$17,6,FALSE),IF(F8>=Master_DB!$B$11,VLOOKUP(C6,Master_DB!$A$16:$H$17,5,FALSE),IF(F8>=Master_DB!$B$10,VLOOKUP(C6,Master_DB!$A$16:$H$17,4,FALSE),VLOOKUP(C6,Master_DB!$A$16:$H$17,3,FALSE))))),"")',
            "0.00",
            False,
        ),
        (11, "Air MIN ($)", '=IFERROR(IF(C6="","",VLOOKUP(C6,Master_DB!$A$16:$H$17,2,FALSE)),"")', "0.00", False),
        (12, "FSC ($/kg)", '=IFERROR(IF(C6="","",VLOOKUP(C6,Master_DB!$A$16:$H$17,7,FALSE)),"")', "0.00", False),
        (13, "SSC ($/kg)", '=IFERROR(IF(C6="","",VLOOKUP(C6,Master_DB!$A$16:$H$17,8,FALSE)),"")', "0.00", False),
    ):
        label_cell(ws[f"E{r}"], lab)
        value_cell(ws[f"F{r}"], formula, fmt, emphasize=emp)
        ws.row_dimensions[r].height = 24

    # 3. Total card
    header_bar(ws, "H5:I5", "TOTAL APPX. AMOUNT", ORANGE)
    ws.merge_cells("H6:I8")
    ws["H6"] = '=IFERROR(IF(H17="","",SUM(H17:H22)),"")'
    ws["H6"].font = fnt(bold=True, size=26)
    ws["H6"].fill = fill(INPUT_BG)
    ws["H6"].alignment = Alignment(horizontal="center", vertical="center")
    ws["H6"].number_format = '"$"#,##0.00'
    paint_merge(ws, "H6:I8", INPUT_BG)
    lock(ws["H6"])

    for r, lab, val, fmt in (
        (10, "Currency", "USD", None),
        (11, "Route", '=IFERROR(C6,"")', None),
        (12, "C.W.", '=IFERROR(F8,"")', "0.00"),
    ):
        label_cell(ws[f"H{r}"], lab)
        ws[f"I{r}"] = val
        ws[f"I{r}"].fill = fill(LIGHT)
        ws[f"I{r}"].border = thin
        ws[f"I{r}"].font = fnt(bold=True, size=11)
        ws[f"I{r}"].alignment = Alignment(horizontal="center", vertical="center")
        if fmt:
            ws[f"I{r}"].number_format = fmt
        lock(ws[f"I{r}"])

    # 4. Breakdown — Item | Unit | Amount only (no formula captions in cells)
    header_bar(ws, "B15:I15", "3. 비용 내역 (Breakdown)", NAVY)
    for col, text, merge in (
        ("B16", "항목", "B16:C16"),
        ("E16", "단위", "E16:F16"),
        ("H16", "금액 (USD)", "H16:I16"),
    ):
        ws.merge_cells(merge)
        ws[col] = text
        paint_merge(ws, merge, SLATE)
        ws[col].font = fnt(bold=True, size=10, color=WHITE)
        ws[col].alignment = Alignment(horizontal="center", vertical="center")

    rows = [
        (
            17,
            "Air Freight",
            '=IFERROR(F9,"")',
            '=IFERROR(IF(OR(F10="",F8="",F11=""),"",MAX(F10*F8,F11)),"")',
        ),
        (18, "FSC", "Per KG", '=IFERROR(IF(OR(F12="",F8=""),"",F12*F8),"")'),
        (19, "SSC", "Per KG", '=IFERROR(IF(OR(F13="",F8=""),"",F13*F8),"")'),
        (
            20,
            "Handling Fee",
            '=IFERROR(INDEX(Master_DB!$B$21:$B$23,MATCH("Handling Fee",Master_DB!$A$21:$A$23,0)),"")',
            '=IFERROR(MAX(INDEX(Master_DB!$C$21:$C$23,MATCH("Handling Fee",Master_DB!$A$21:$A$23,0)),INDEX(Master_DB!$D$21:$D$23,MATCH("Handling Fee",Master_DB!$A$21:$A$23,0))),"")',
        ),
        (
            21,
            "Doc Fee",
            '=IFERROR(INDEX(Master_DB!$B$21:$B$23,MATCH("Doc Fee",Master_DB!$A$21:$A$23,0)),"")',
            '=IFERROR(IF(C12="","",MAX(INDEX(Master_DB!$C$21:$C$23,MATCH("Doc Fee",Master_DB!$A$21:$A$23,0)),INDEX(Master_DB!$D$21:$D$23,MATCH("Doc Fee",Master_DB!$A$21:$A$23,0)))*C12),"")',
        ),
        (
            22,
            "Trucking",
            '=IFERROR(INDEX(Master_DB!$B$21:$B$23,MATCH("Trucking",Master_DB!$A$21:$A$23,0)),"")',
            '=IFERROR(IF(F6="","",MAX(INDEX(Master_DB!$C$21:$C$23,MATCH("Trucking",Master_DB!$A$21:$A$23,0))*F6,INDEX(Master_DB!$D$21:$D$23,MATCH("Trucking",Master_DB!$A$21:$A$23,0)))),"")',
        ),
    ]

    for r, item, unit, amt in rows:
        ws.merge_cells(f"B{r}:C{r}")
        ws[f"B{r}"] = item
        ws[f"B{r}"].font = fnt(bold=True, size=11)
        ws[f"B{r}"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
        paint_merge(ws, f"B{r}:C{r}", WHITE)

        ws.merge_cells(f"E{r}:F{r}")
        ws[f"E{r}"] = unit
        ws[f"E{r}"].font = fnt(size=10, color=MUTED)
        ws[f"E{r}"].alignment = Alignment(horizontal="center", vertical="center")
        paint_merge(ws, f"E{r}:F{r}", LIGHT)

        ws.merge_cells(f"H{r}:I{r}")
        ws[f"H{r}"] = amt
        ws[f"H{r}"].number_format = '"$"#,##0.00'
        ws[f"H{r}"].font = fnt(bold=True, size=11)
        ws[f"H{r}"].alignment = Alignment(horizontal="center", vertical="center")
        paint_merge(ws, f"H{r}:I{r}", LIGHT)
        ws.row_dimensions[r].height = 26

    ws["B24"] = "CASE 입력값·기대결과는 Guide 시트  |  요율 변경은 Master_DB 노란 칸만"
    ws["B24"].font = fnt(size=9, color=MUTED)
    ws.merge_cells("B24:I24")
    lock(ws["B24"])

    ws.protection.sheet = True
    ws.protection.enable()


def build_guide(ws):
    """Submission-ready brief for CM (no tutorial tone)."""
    ws.sheet_view.showGridLines = False
    set_widths(
        ws,
        {
            "A": 2,
            "B": 12,
            "C": 8,
            "D": 8,
            "E": 8,
            "F": 8,
            "G": 10,
            "H": 14,
            "I": 26,
            "J": 12,
        },
    )

    ws["B2"] = "작업 안내"
    ws["B2"].font = fnt(bold=True, size=16)

    ws["B4"] = "1. 파일 구성"
    ws["B4"].font = fnt(bold=True, size=12, color=ORANGE)
    for i, t in enumerate(
        [
            "Master_DB : 항공 구간 요율(Weight Break)·할증(FSC/SSC)·로컬 단가 관리 (Backend)",
            "Quote : 화물 입력 및 가견적 산출 화면 (Frontend / 원페이지 대시보드)",
            "요율 변경 시 Master_DB의 노란색 셀만 수정. Quote 수식에는 단가를 직접 기입하지 않음.",
        ]
    ):
        cell = ws.cell(5 + i, 2, t)
        cell.font = fnt(size=11)
        ws.merge_cells(start_row=5 + i, start_column=2, end_row=5 + i, end_column=10)

    ws["B9"] = "2. 사용 방법"
    ws["B9"].font = fnt(bold=True, size=12, color=ORANGE)
    for i, t in enumerate(
        [
            "Quote 시트에서 Route, 치수(L/W/H), 수량, 실중량, B/L 수를 입력한다. (주황색 셀)",
            "CBM·부피중량·C.W.·적용 Weight Break·항목별 금액·TOTAL APPX.는 자동 계산된다.",
            "마스터 요율·MIN 변경은 Master_DB에서 처리하며, 필요 시 차장 확인 후 반영한다.",
        ]
    ):
        cell = ws.cell(10 + i, 2, t)
        cell.font = fnt(size=11)
        ws.merge_cells(start_row=10 + i, start_column=2, end_row=10 + i, end_column=10)

    ws["B14"] = "3. 산출 기준 (의뢰서)"
    ws["B14"].font = fnt(bold=True, size=12, color=ORANGE)
    for i, t in enumerate(
        [
            "CBM = L × W × H × Qty ÷ 1,000,000  /  부피중량 = CBM × 167  /  C.W. = MAX(실중량, 부피중량)",
            "항공운임 = MAX(적용단가 × C.W., MIN)  /  FSC·SSC = 각 단가 × C.W.",
            "로컬: Handling·Doc는 건(BL) 단위, Trucking = MAX(단가 × CBM, MIN)",
            "Quote 단가는 Master_DB를 VLOOKUP·INDEX/MATCH로 참조. 빈 입력 시 IFERROR로 오류 표시 방지.",
        ]
    ):
        cell = ws.cell(15 + i, 2, t)
        cell.font = fnt(size=11)
        ws.merge_cells(start_row=15 + i, start_column=2, end_row=15 + i, end_column=10)

    ws["B20"] = "4. 검증 CASE (Route: ICN-HKG) — Quote 입력란에 아래 수치 적용"
    ws["B20"].font = fnt(bold=True, size=12, color=ORANGE)
    ws.merge_cells("B20:J20")

    headers = ["CASE", "L", "W", "H", "Qty", "Gross", "C.W.", "확인 포인트", "TOTAL APPX."]
    for c, h in enumerate(headers, 2):
        cell = ws.cell(21, c, h)
        cell.fill = fill(NAVY)
        cell.font = fnt(bold=True, size=10, color=WHITE)
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cell.border = thin
    ws.row_dimensions[21].height = 28

    for i, row in enumerate(
        [
            ("A 부피화물", 110, 110, 150, 3, 400, 909.32, "Break +500 / $3.20", 3728.47),
            ("B 고중량", 30, 30, 30, 1, 80, 80.00, "Break +45 / $4.50", 555.00),
            ("C 미니멈", 20, 20, 20, 1, 3, 3.00, "Air $50 · Truck $80", 187.25),
        ]
    ):
        r = 22 + i
        for c, v in enumerate(row, 2):
            cell = ws.cell(r, c, v)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = fill(LIGHT)
            if c == 8:
                cell.number_format = "0.00"
            if c == 10:
                cell.number_format = '"$"#,##0.00'

    ws["B26"] = "5. 데이터·배포"
    ws["B26"].font = fnt(bold=True, size=12, color=ORANGE)
    for i, t in enumerate(
        [
            "본 파일의 Master 수치는 의뢰서 기준 가상(dummy) 데이터이다, 검증용 CASE 3건으로 로직을 확인한 상태이다.",
            "실요율 반영 및 항목 추가는 별도 확인 후 Master_DB에 갱신한다. (Quote 화면 구조는 유지)",
            "시트 보호: Quote는 입력 셀만 편집 가능. 수식 수정 시 [검토] → 시트 보호 해제.",
            "클라우드 업로드·권한: 10층 Room H Teddy 담당.",
        ]
    ):
        cell = ws.cell(27 + i, 2, t)
        cell.font = fnt(size=11)
        ws.merge_cells(start_row=27 + i, start_column=2, end_row=27 + i, end_column=10)

    ws["B32"] = "주황=입력  /  노랑(Master)=요율  /  초록=C.W.  /  회색=자동산출"
    ws["B32"].font = fnt(size=9, color=MUTED)
    ws.merge_cells("B32:J32")


def verify_logic():
    vol_factor, divisor = 167, 1_000_000
    air = {"MIN": 50, 45: 4.5, 100: 3.8, 500: 3.2, 1000: 2.8, "FSC": 0.6, "SSC": 0.15}

    def one(L, W, H, qty, gross):
        cbm = L * W * H * qty / divisor
        cw = max(gross, cbm * vol_factor)
        if cw >= 1000:
            br, rate = 1000, air[1000]
        elif cw >= 500:
            br, rate = 500, air[500]
        elif cw >= 100:
            br, rate = 100, air[100]
        else:
            br, rate = 45, air[45]
        air_amt = max(rate * cw, air["MIN"])
        total = air_amt + air["FSC"] * cw + air["SSC"] * cw + 30 + 25 + max(15 * cbm, 80)
        return {"cw": cw, "br": br, "rate": rate, "air": air_amt, "truck": max(15 * cbm, 80), "total": total}

    a, b, c = one(110, 110, 150, 3, 400), one(30, 30, 30, 1, 80), one(20, 20, 20, 1, 3)
    assert abs(a["cw"] - 909.315) < 0.01 and a["br"] == 500
    assert b["cw"] == 80 and b["rate"] == 4.5
    assert c["air"] == 50 and c["truck"] == 80
    return {"A": a, "B": b, "C": c}


def main():
    cases = verify_logic()
    print("Logic OK")
    for k, v in cases.items():
        print(f"  {k}: CW={v['cw']:.2f} +{v['br']} @{v['rate']} TOTAL={v['total']:.2f}")

    wb = Workbook()
    ws_q = wb.active
    ws_q.title = "Quote"
    ws_m = wb.create_sheet("Master_DB", 0)
    ws_g = wb.create_sheet("Guide", 2)
    build_master(ws_m)
    build_quote(ws_q)
    build_guide(ws_g)
    wb._sheets = [ws_m, ws_q, ws_g]
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
